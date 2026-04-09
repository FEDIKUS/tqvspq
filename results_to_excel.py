#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import pathlib
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_PATTERNS = ("output-*.json",)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export benchmark output JSON files to an Excel workbook."
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        help="JSON files or glob patterns. Default: output-*.json",
    )
    parser.add_argument(
        "--output",
        default="benchmark_results.xlsx",
        help="Output .xlsx path",
    )
    return parser.parse_args()


def expand_inputs(inputs: list[str]) -> list[pathlib.Path]:
    patterns = inputs or list(DEFAULT_PATTERNS)
    paths: list[pathlib.Path] = []
    seen: set[pathlib.Path] = set()
    for pattern in patterns:
        matches = sorted(pathlib.Path().glob(pattern))
        if matches:
            for path in matches:
                resolved = path.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    paths.append(path)
            continue

        path = pathlib.Path(pattern)
        if path.is_file():
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                paths.append(path)

    if not paths:
        joined = ", ".join(patterns)
        raise FileNotFoundError(f"No JSON files matched: {joined}")
    return paths


def load_json(path: pathlib.Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def flatten_run(path: pathlib.Path, payload: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    config = payload.get("config", {})
    run_row = {
        "source_file": path.name,
        "started_at_utc": payload.get("started_at_utc"),
        "finished_at_utc": payload.get("finished_at_utc"),
        "methods": ",".join(config.get("methods", [])) if isinstance(config.get("methods"), list) else config.get("methods"),
        "pq_bits": config.get("pq_bits"),
        "tq_bits": config.get("tq_bits"),
        "tq_backend": config.get("tq_backend"),
        "topk": config.get("topk"),
        "pq_train_size": config.get("pq_train_size"),
        "base_batch_size": config.get("base_batch_size"),
        "query_batch_size": config.get("query_batch_size"),
        "workers": config.get("workers"),
        "faiss_threads": config.get("faiss_threads"),
        "torch_threads": config.get("torch_threads"),
        "torch_interop_threads": config.get("torch_interop_threads"),
        "seed": config.get("seed"),
    }

    result_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for result in payload.get("results", []):
        base_result = {
            "source_file": path.name,
            "dataset": result.get("dataset"),
            "base_count": result.get("base_count"),
            "query_count": result.get("query_count"),
            "dimension": result.get("dimension"),
            "requested_topk": result.get("requested_topk", result.get("topk")),
            "topk": result.get("topk"),
            "topk_source": result.get("topk_source"),
            "topk_reference_seconds": result.get("topk_reference_seconds"),
        }

        if "error" in result:
            error_rows.append(
                {
                    **base_result,
                    "error": result.get("error"),
                    "traceback": result.get("traceback"),
                }
            )
            continue

        methods = result.get("methods", {})
        for method_name, method_payload in methods.items():
            recon = method_payload.get("reconstruction_l2", {})
            drift = method_payload.get("topk_distance_change_l2", {})
            result_rows.append(
                {
                    **base_result,
                    "method": method_name,
                    "backend": method_payload.get("backend"),
                    "bits_per_coordinate": method_payload.get("bits_per_coordinate"),
                    "subspace_size": method_payload.get("subspace_size"),
                    "bytes_per_vector": method_payload.get("bytes_per_vector"),
                    "packed_index_bytes_per_vector": method_payload.get("packed_index_bytes_per_vector"),
                    "norm_bytes_per_vector": method_payload.get("norm_bytes_per_vector"),
                    "train_seconds": method_payload.get("train_seconds"),
                    "reconstruction_seconds": method_payload.get("reconstruction_seconds"),
                    "neighbor_predecode_seconds": method_payload.get("neighbor_predecode_seconds"),
                    "topk_distance_seconds": method_payload.get("topk_distance_seconds"),
                    "total_seconds": method_payload.get("total_seconds"),
                    "recon_mean": recon.get("mean"),
                    "recon_std": recon.get("std"),
                    "recon_rms": recon.get("rms"),
                    "recon_max": recon.get("max"),
                    "topk_mean_signed_change": drift.get("mean_signed_change"),
                    "topk_mean_absolute_change": drift.get("mean_absolute_change"),
                    "topk_rmse_change": drift.get("rmse_change"),
                    "topk_max_absolute_change": drift.get("max_absolute_change"),
                }
            )

    return run_row, result_rows, error_rows


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet.append(["empty"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    float_headers = {
        "topk_reference_seconds",
        "train_seconds",
        "reconstruction_seconds",
        "neighbor_predecode_seconds",
        "topk_distance_seconds",
        "total_seconds",
        "recon_mean",
        "recon_std",
        "recon_rms",
        "recon_max",
        "topk_mean_signed_change",
        "topk_mean_absolute_change",
        "topk_rmse_change",
        "topk_max_absolute_change",
    }

    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = len(header)
        for cell in worksheet[column_letter]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
            if header in float_headers and cell.row > 1 and isinstance(value, (int, float)):
                cell.number_format = "0.000000"
        worksheet.column_dimensions[column_letter].width = min(max_len + 2, 60)


def main() -> int:
    args = parse_args()
    input_paths = expand_inputs(args.inputs)

    run_rows: list[dict[str, Any]] = []
    result_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for path in input_paths:
        payload = load_json(path)
        run_row, rows, errors = flatten_run(path, payload)
        run_rows.append(run_row)
        result_rows.extend(rows)
        error_rows.extend(errors)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(workbook, "results", result_rows)
    write_sheet(workbook, "runs", run_rows)
    write_sheet(workbook, "errors", error_rows)

    output_path = pathlib.Path(args.output)
    workbook.save(output_path)
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
