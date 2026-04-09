#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import pathlib
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEFAULT_PATTERNS = ("output-top*.json",)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export ann_benchmark2 output JSON files to an Excel workbook."
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        help="JSON files or glob patterns. Default: output-top*.json",
    )
    parser.add_argument(
        "--output",
        default="benchmark_top_results.xlsx",
        help="Output .xlsx path",
    )
    return parser.parse_args()


def expand_inputs(inputs: list[str]) -> list[pathlib.Path]:
    patterns = inputs or list(DEFAULT_PATTERNS)
    paths: list[pathlib.Path] = []
    seen: set[pathlib.Path] = set()
    for pattern in patterns:
        matches = sorted(pathlib.Path().glob(pattern))
        if matches:
            for path in matches:
                resolved = path.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    paths.append(path)
            continue

        path = pathlib.Path(pattern)
        if path.is_file():
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                paths.append(path)

    if not paths:
        raise FileNotFoundError(f"No JSON files matched: {', '.join(patterns)}")
    return paths


def load_json(path: pathlib.Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def flatten_run(
    path: pathlib.Path, payload: dict[str, Any]
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    config = payload.get("config", {})
    run_row = {
        "source_file": path.name,
        "method": config.get("method"),
        "candidate_topx": config.get("candidate_topx"),
        "compare_topk": config.get("compare_topk"),
        "pq_bits": config.get("pq_bits"),
        "tq_bits": config.get("tq_bits"),
        "pq_train_size": config.get("pq_train_size"),
        "base_batch_size": config.get("base_batch_size"),
        "query_batch_size": config.get("query_batch_size"),
        "faiss_threads": config.get("faiss_threads"),
        "torch_threads": config.get("torch_threads"),
        "torch_interop_threads": config.get("torch_interop_threads"),
        "seed": config.get("seed"),
    }

    result_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for result in payload.get("results", []):
        base_row = {
            "source_file": path.name,
            "dataset": result.get("dataset"),
            "base_count": result.get("base_count"),
            "query_count": result.get("query_count"),
            "dimension": result.get("dimension"),
            "candidate_topx_requested": result.get("candidate_topx_requested"),
            "candidate_topx_used": result.get("candidate_topx_used"),
            "compare_topk_requested": result.get("compare_topk_requested"),
            "compare_topk_used": result.get("compare_topk_used"),
        }

        if "error" in result:
            error_rows.append(
                {
                    **base_row,
                    "error": result.get("error"),
                }
            )
            continue

        result_rows.append(
            {
                **base_row,
                "method": result.get("method"),
                "bits_per_coordinate": result.get("bits_per_coordinate"),
                "bytes_per_vector": result.get("bytes_per_vector"),
                "packed_index_bytes_per_vector": result.get("packed_index_bytes_per_vector"),
                "norm_bytes_per_vector": result.get("norm_bytes_per_vector"),
                "subspace_size": result.get("subspace_size"),
                "ground_truth_load_seconds": result.get("ground_truth_load_seconds"),
                "train_seconds": result.get("train_seconds"),
                "decode_candidates_seconds": result.get("decode_candidates_seconds"),
                "rerank_seconds": result.get("rerank_seconds"),
                "total_seconds": result.get("total_seconds"),
                "recall_percent": result.get("recall_percent"),
                "full_recovery_percent": result.get("full_recovery_percent"),
                "exact_order_match_percent": result.get("exact_order_match_percent"),
            }
        )

    return run_row, result_rows, error_rows


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet.append(["empty"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    float_headers = {
        "ground_truth_load_seconds",
        "train_seconds",
        "decode_candidates_seconds",
        "rerank_seconds",
        "total_seconds",
        "recall_percent",
        "full_recovery_percent",
        "exact_order_match_percent",
    }

    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = len(header)
        for cell in worksheet[column_letter]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
            if header in float_headers and cell.row > 1 and isinstance(value, (int, float)):
                cell.number_format = "0.000000"
        worksheet.column_dimensions[column_letter].width = min(max_len + 2, 60)


def main() -> int:
    args = parse_args()
    input_paths = expand_inputs(args.inputs)

    run_rows: list[dict[str, Any]] = []
    result_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for path in input_paths:
        payload = load_json(path)
        run_row, rows, errors = flatten_run(path, payload)
        run_rows.append(run_row)
        result_rows.extend(rows)
        error_rows.extend(errors)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(workbook, "results", result_rows)
    write_sheet(workbook, "runs", run_rows)
    write_sheet(workbook, "errors", error_rows)

    output_path = pathlib.Path(args.output)
    workbook.save(output_path)
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
